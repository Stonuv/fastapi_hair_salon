from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from ..repositories.report_repository import ReportRepository
from ..schemas.report import DailyRevenue, MasterReportRow, ReportResponse, ServiceReportRow

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="78350F")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_table_header(ws, row: int, columns: list[str]) -> None:
    for col, label in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


class ReportService:
    def __init__(self, db: Session):
        self._repo = ReportRepository(db)

    def get_report(self, date_from: date, date_to: date) -> ReportResponse:
        count, revenue, avg = self._repo.get_summary(date_from, date_to)
        repeat_pct = self._repo.get_repeat_clients_pct(date_from, date_to)

        # Денежные значения из SQL приходят как Decimal и остаются Decimal
        # до самой сериализации (float терял бы точность).
        revenue_by_day = [
            DailyRevenue(date=row.date, revenue=row.revenue)
            for row in self._repo.get_revenue_by_day(date_from, date_to)
        ]
        by_service = [
            ServiceReportRow(service_name=row.service_name, appointments=row.appointments)
            for row in self._repo.get_appointments_by_service(date_from, date_to)
        ]
        masters = [
            MasterReportRow(
                master_name=row.master_name,
                appointments=row.appointments,
                revenue=row.revenue,
                avg_check=row.avg_check,
                avg_rating=round(float(row.avg_rating), 2) if row.avg_rating is not None else None,
            )
            for row in self._repo.get_masters_breakdown(date_from, date_to)
        ]
        return ReportResponse(
            date_from=date_from,
            date_to=date_to,
            total_revenue=revenue,
            total_appointments=count,
            avg_check=avg,
            repeat_clients_pct=repeat_pct,
            revenue_by_day=revenue_by_day,
            appointments_by_service=by_service,
            masters_breakdown=masters,
        )

    def export_excel(self, date_from: date, date_to: date) -> bytes:
        report = self.get_report(date_from, date_to)
        wb = openpyxl.Workbook()

        # ── Сводка ──────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Сводка"
        ws["A1"] = f"Отчёт за период: {date_from} — {date_to}"
        ws["A1"].font = Font(bold=True, size=13)
        ws.append([])
        _write_table_header(ws, 3, ["Показатель", "Значение"])
        rows = [
            ("Общая выручка (₽)", report.total_revenue),
            ("Всего записей", report.total_appointments),
            ("Средний чек (₽)", round(report.avg_check, 2)),
            ("Повторные клиенты (%)", report.repeat_clients_pct),
        ]
        for i, (label, value) in enumerate(rows, start=4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

        # ── Выручка по дням ─────────────────────────────────────────────
        ws2 = wb.create_sheet("Выручка по дням")
        _write_table_header(ws2, 1, ["Дата", "Выручка (₽)"])
        for row in report.revenue_by_day:
            ws2.append([str(row.date), row.revenue])
        ws2.column_dimensions["A"].width = 14
        ws2.column_dimensions["B"].width = 16

        # ── По услугам ──────────────────────────────────────────────────
        ws3 = wb.create_sheet("По услугам")
        _write_table_header(ws3, 1, ["Услуга", "Записей"])
        for row in report.appointments_by_service:
            ws3.append([row.service_name, row.appointments])
        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 12

        # ── Мастера ─────────────────────────────────────────────────────
        ws4 = wb.create_sheet("Мастера")
        _write_table_header(ws4, 1, ["Мастер", "Записей", "Выручка (₽)", "Средний чек (₽)", "Рейтинг"])
        for row in report.masters_breakdown:
            ws4.append([
                row.master_name,
                row.appointments,
                row.revenue,
                round(row.avg_check, 2),
                row.avg_rating if row.avg_rating is not None else "—",
            ])
        for col_letter, width in zip(["A", "B", "C", "D", "E"], [25, 10, 16, 18, 10], strict=True):
            ws4.column_dimensions[col_letter].width = width

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
